import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection, NamedStyle
from openpyxl.styles.colors import Color
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
import os
from openpyxl.formatting.rule import FormulaRule
import re
from openpyxl.workbook.defined_name import DefinedName



def generate_code_formula(prefix, start_row):
    """יצירת נוסחה לקוד אוטומטי"""
    return f'=IF(B{start_row}="","","{prefix}"&TEXT(ROW()-1,"000"))'

def generate_subcat_code_formula(start_row):
    """יצירת נוסחה לקוד קטגוריית משנה"""
    return f'=IF(AND(B{start_row}<>"",D{start_row}<>""),INDEX(\'קטגוריות ראשיות\'!$A$2:$A$1000,MATCH(D{start_row},\'קטגוריות ראשיות\'!$B$2:$B$1000,0))&"-"&TEXT(COUNTIF(D$2:D{start_row},D{start_row}),"000"),"")'

def create_master_template(lang='he'):
    """
    יצירת תבנית אקסל
    lang: 'he' לעברית, 'en' לאנגלית
    """
    VERSION = "1.0.0"  # גרסת התבנית
    CURRENT_DATE = "30/03/2025"  # תאריך עדכון אחרון
    SHEET_PASSWORD = "FlexiPI@2025#System"  # סיסמה להגנת הגליון
    SUPPORT_EMAIL = "support@flexipi.com"  # כתובת אימייל לתמיכה
    
    # מילון תרגומים
    translations = {
        'he': {
            'direction': 'rtl',
            'alignment': 'right',
            'sheets': {
                'main_categories': 'קטגוריות ראשיות',
                'sub_categories': 'קטגוריות משנה',
                'colors': 'צבעים',
                'sizes': 'מידות',
                'materials': 'חומרים',
                'suppliers': 'ספקים',
                'products': 'מוצרים',
                'instructions': 'הנחיות',
                'settings': 'הגדרות'
            },
            'headers': {
                'code': 'קוד',
                'name': 'שם',
                'description': 'תיאור',
                'status': 'סטטוס',
                'action': 'פעולה',
                'main_category': 'קטגוריה ראשית',
                'sub_category': 'קטגוריית משנה',
                'contact': 'איש קשר',
                'email': 'אימייל',
                'phone': 'טלפון',
                'address': 'כתובת',
                'sku': 'מק"ט',
                'supplier': 'ספק',
                'colors': 'צבעים',
                'sizes': 'מידות',
                'materials': 'חומרים',  # שונה מ'חומר' ל'חומרים'
                'units_per_pack': 'כמות באריזה',
                'packing_instructions': 'הוראות אריזה',
                'units_per_carton': 'כמות בקרטון',
                'price_per_unit': 'מחיר ליחידה'
            },
            'validations': {
                'status': ['פעיל', 'לא פעיל'],
                'action': ['עדכון', 'מחיקה', 'הוספה'],
                'error_title': 'שגיאה',
                'error_message': 'יש לבחור ערך מהרשימה בלבד'
            },
            'instructions_text': [
                ['הנחיות שימוש בקובץ טמפלייט לניהול הזמנות FlexiPI'],
                [''],
                [f'גרסה: {VERSION}'],
                [f'תאריך עדכון אחרון: {CURRENT_DATE}'],
                ['כל הזכויות שמורות © FlexiPI'],
                [''],
                ['מבנה הקובץ:'],
                ['הקובץ מכיל את הגיליונות הבאים:'],
                ['הנחיות - הגיליון הנוכחי .1'],
                ['הגדרות - הגדרות מערכת כגון מבנה מק"ט .2'],
                ['קטגוריות ראשיות - ניהול קטגוריות ראשיות של מוצרים .3'],
                ['קטגוריות משנה - ניהול תתי-קטגוריות .4'],
                ['צבעים - ניהול צבעים זמינים .5'],
                ['מידות - ניהול מידות זמינות .6'],
                ['חומרים - ניהול חומרים זמינים .7'],
                ['ספקים - ניהול פרטי ספקים .8'],
                ['מוצרים - ניהול פריטי המלאי .9'],
                [''],
                ['הנחיות כלליות:'],
                ['יש למלא את הנתונים בסדר הבא: .1'],
                ['קטגוריות ראשיות .א'],
                ['קטגוריות משנה .ב'],
                ['צבעים .ג'],
                ['מידות .ד'],
                ['חומרים .ה'],
                ['ספקים .ו'],
                ['מוצרים .ז'],
                [''],
                ['עמודות הקוד והמק"ט נעולות ומתמלאות אוטומטית .2'],
                ['בשדות עם רשימה נפתחת יש לבחור מהרשימה בלבד .3'],
                ['אין לשנות את מבנה הקובץ או להוסיף/למחוק עמודות .4'],
                ['ניתן להוסיף שורות נתונים ככל שנדרש .5'],
                [''],
                ['פירוט השדות בכל גיליון:'],
                [''],
                ['קטגוריות ראשיות:'],
                ['קוד: נוצר אוטומטית (CAT001, CAT002, ...) -'],
                ['שם: שם הקטגוריה (לדוגמה: "כלי מטבח", "ריהוט") -'],
                ['תיאור: תיאור מפורט של הקטגוריה -'],
                ['סטטוס: פעיל/לא פעיל -'],
                ['פעולה: עדכון/מחיקה/הוספה -'],
                [''],
                ['קטגוריות משנה:'],
                ['קוד: נוצר אוטומטית (CAT001-001, CAT001-002, ...) -'],
                ['שם: שם תת-הקטגוריה (לדוגמה: "סכו"ם", "כיסאות") -'],
                ['תיאור: תיאור מפורט של תת-הקטגוריה -'],
                ['קטגוריה ראשית: בחירה מרשימה נפתחת -'],
                ['סטטוס: פעיל/לא פעיל -'],
                ['פעולה: עדכון/מחיקה/הוספה -'],
                [''],
                ['צבעים:'],
                ['קוד: נוצר אוטומטית (COL001, COL002, ...) -'],
                ['שם: שם הצבע (לדוגמה: "אדום", "כחול") -'],
                ['תיאור: תיאור מפורט או קוד צבע -'],
                ['סטטוס: פעיל/לא פעיל -'],
                ['פעולה: עדכון/מחיקה/הוספה -'],
                [''],
                ['מידות:'],
                ['קוד: נוצר אוטומטית (SIZ001, SIZ002, ...) -'],
                ['שם: המידה (לדוגמה: "S", "M", "L", "XL") -'],
                ['תיאור: פירוט מדויק של המידה -'],
                ['סטטוס: פעיל/לא פעיל -'],
                ['פעולה: עדכון/מחיקה/הוספה -'],
                [''],
                ['חומרים:'],
                ['קוד: נוצר אוטומטית (MAT001, MAT002, ...) -'],
                ['שם: שם החומר (לדוגמה: "עץ אלון", "פלסטיק") -'],
                ['תיאור: מפרט טכני של החומר -'],
                ['סטטוס: פעיל/לא פעיל -'],
                ['פעולה: עדכון/מחיקה/הוספה -'],
                [''],
                ['ספקים:'],
                ['קוד: נוצר אוטומטית (SUP001, SUP002, ...) -'],
                ['שם: שם החברה -'],
                ['איש קשר: שם איש הקשר -'],
                ['אימייל: כתובת דוא"ל -'],
                ['טלפון: מספר טלפון -'],
                ['כתובת: כתובת מלאה -'],
                ['סטטוס: פעיל/לא פעיל -'],
                ['פעולה: עדכון/מחיקה/הוספה -'],
                [''],
                ['מוצרים:'],
                ['מק"ט: נוצר אוטומטית לפי הגדרות -'],
                ['שם: שם המוצר -'],
                ['תיאור: תיאור מפורט של המוצר -'],
                ['קטגוריה ראשית: בחירה מרשימה -'],
                ['קטגוריית משנה: בחירה מרשימה -'],
                ['ספק: בחירה מרשימה -'],
                ['צבעים: בחירה מרשימה -'],
                ['מידות: בחירה מרשימה -'],
                ['חומרים: בחירה מרשימה -'],
                ['כמות באריזה: מספר יחידות -'],
                ['הוראות אריזה: הנחיות מיוחדות -'],
                ['כמות בקרטון: מספר יחידות -'],
                ['מחיר ליחידה: מחיר ב-USD -'],
                ['סטטוס: פעיל/לא פעיל -'],
                ['פעולה: עדכון/מחיקה/הוספה -'],
                [''],
                ['הערות חשובות:'],
                ['יש לוודא שכל השדות החובה מלאים .1'],
                ['בעת הוספת מוצר חדש, יש לוודא שכל הערכים הנבחרים (קטגוריות, צבעים וכו\') קיימים ופעילים .2'],
                ['מחירים יש להזין במספרים בלבד, ללא סימנים מיוחדים .3'],
                ['בכל שינוי יש לציין את סוג הפעולה בעמודת "פעולה" .4'],
                ['אין למחוק שורות קיימות, אלא רק לסמן אותן כ"לא פעיל" .5'],
                [''],
                ['תמיכה:'],
                ['במקרה של תקלה או שאלה, ניתן לפנות לתמיכה הטכנית:'],
                [SUPPORT_EMAIL]
            ]
        },
        'en': {
            'direction': 'ltr',
            'alignment': 'left',
            'sheets': {
                'main_categories': 'Main Categories',
                'sub_categories': 'Sub Categories',
                'colors': 'Colors',
                'sizes': 'Sizes',
                'materials': 'Materials',
                'suppliers': 'Suppliers',
                'products': 'Products',
                'instructions': 'Instructions',
                'settings': 'Settings'
            },
            'headers': {
                'code': 'Code',
                'name': 'Name',
                'description': 'Description',
                'status': 'Status',
                'action': 'Action',
                'main_category': 'Main Category',
                'sub_category': 'Sub Category',
                'contact': 'Contact',
                'email': 'Email',
                'phone': 'Phone',
                'address': 'Address',
                'sku': 'SKU',
                'supplier': 'Supplier',
                'colors': 'Colors',
                'sizes': 'Sizes',
                'materials': 'Materials',  # Changed from 'Material' to 'Materials'
                'units_per_pack': 'Units Per Pack',
                'packing_instructions': 'Packing Instructions',
                'units_per_carton': 'Units Per Carton',
                'price_per_unit': 'Price Per Unit'
            },
            'validations': {
                'status': ['Active', 'Inactive'],
                'action': ['Update', 'Delete', 'Add'],
                'error_title': 'Error',
                'error_message': 'Please select a value from the list'
            },
            'instructions_text': [
                ['FlexiPI Template Usage Instructions'],
                [''],
                [f'Version: {VERSION}'],
                [f'Last Updated: {CURRENT_DATE}'],
                ['All Rights Reserved © FlexiPI'],
                [''],
                ['File Structure:'],
                ['The file contains the following sheets:'],
                ['1. Instructions - Current sheet'],
                ['2. Settings - System settings like SKU structure'],
                ['3. Main Categories - Manage main product categories'],
                ['4. Sub Categories - Manage sub-categories'],
                ['5. Colors - Manage available colors'],
                ['6. Sizes - Manage available sizes'],
                ['7. Materials - Manage available materials'],
                ['8. Suppliers - Manage supplier details'],
                ['9. Products - Manage inventory items'],
                [''],
                ['General Instructions:'],
                ['1. Fill in the data in the following order:'],
                ['   a. Main Categories'],
                ['   b. Sub Categories'],
                ['   c. Colors'],
                ['   d. Sizes'],
                ['   e. Materials'],
                ['   f. Suppliers'],
                ['   g. Products'],
                [''],
                ['2. Code and SKU columns are locked and auto-generated'],
                ['3. For fields with dropdown lists, select from the list only'],
                ['4. Do not modify the file structure or add/delete columns'],
                ['5. You can add data rows as needed'],
                [''],
                ['Field Details for Each Sheet:'],
                [''],
                ['Main Categories:'],
                ['- Code: Auto-generated (CAT001, CAT002, ...)'],
                ['- Name: Category name (e.g., "Kitchen Utensils", "Furniture")'],
                ['- Description: Detailed category description'],
                ['- Status: Active/Inactive'],
                ['- Action: Update/Delete/Add'],
                [''],
                ['Sub Categories:'],
                ['- Code: Auto-generated (CAT001-001, CAT001-002, ...)'],
                ['- Name: Sub-category name (e.g., "Cutlery", "Chairs")'],
                ['- Description: Detailed sub-category description'],
                ['- Main Category: Select from dropdown'],
                ['- Status: Active/Inactive'],
                ['- Action: Update/Delete/Add'],
                [''],
                ['Colors:'],
                ['- Code: Auto-generated (COL001, COL002, ...)'],
                ['- Name: Color name (e.g., "Red", "Blue")'],
                ['- Description: Detailed description or color code'],
                ['- Status: Active/Inactive'],
                ['- Action: Update/Delete/Add'],
                [''],
                ['Sizes:'],
                ['- Code: Auto-generated (SIZ001, SIZ002, ...)'],
                ['- Name: Size (e.g., "S", "M", "L", "XL")'],
                ['- Description: Exact size specifications'],
                ['- Status: Active/Inactive'],
                ['- Action: Update/Delete/Add'],
                [''],
                ['Materials:'],
                ['- Code: Auto-generated (MAT001, MAT002, ...)'],
                ['- Name: Material name (e.g., "Oak Wood", "Plastic")'],
                ['- Description: Technical specifications'],
                ['- Status: Active/Inactive'],
                ['- Action: Update/Delete/Add'],
                [''],
                ['Suppliers:'],
                ['- Code: Auto-generated (SUP001, SUP002, ...)'],
                ['- Name: Company name'],
                ['- Contact: Contact person name'],
                ['- Email: Email address'],
                ['- Phone: Phone number'],
                ['- Address: Full address'],
                ['- Status: Active/Inactive'],
                ['- Action: Update/Delete/Add'],
                [''],
                ['Products:'],
                ['- SKU: Auto-generated based on settings'],
                ['- Name: Product name'],
                ['- Description: Detailed product description'],
                ['- Main Category: Select from list'],
                ['- Sub Category: Select from list'],
                ['- Supplier: Select from list'],
                ['- Colors: Select from list'],
                ['- Sizes: Select from list'],
                ['- Materials: Select from list'],
                ['- Units Per Pack: Number of units'],
                ['- Packing Instructions: Special instructions'],
                ['- Units Per Carton: Number of units'],
                ['- Price Per Unit: Price in USD'],
                ['- Status: Active/Inactive'],
                ['- Action: Update/Delete/Add'],
                [''],
                ['Important Notes:'],
                ['1. Ensure all required fields are filled'],
                ['2. When adding a new product, verify all selected values (categories, colors, etc.) exist and are active'],
                ['3. Enter prices as numbers only, without special characters'],
                ['4. Specify the action type in the "Action" column for any changes'],
                ['5. Do not delete existing rows, mark them as "Inactive" instead'],
                [''],
                ['Support:'],
                ['For technical support or questions, contact:'],
                [SUPPORT_EMAIL]
            ]
        }
    }

    t = translations[lang]
    
    # יצירת תיקיית היעד אם לא קיימת
    output_dir = os.path.join('public', 'templates')
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f'master_template_{lang}_v4.xlsx')

    # יצירת חוברת עבודה חדשה
    wb = openpyxl.Workbook()
    
    # ========== יצירת סגנונות כלליים ==========
    
    # סגנון כותרת
    header_style = NamedStyle(name='header_style')
    header_style.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_style.fill = PatternFill(fill_type='solid', start_color='2E7D32', end_color='2E7D32')
    header_style.border = Border(
        left=Side(border_style='medium', color='000000'),
        right=Side(border_style='medium', color='000000'),
        top=Side(border_style='medium', color='000000'),
        bottom=Side(border_style='medium', color='000000')
    )
    header_style.alignment = Alignment(
        horizontal='center',  # משאיר את הכותרות ממורכזות
        vertical='center', 
        wrap_text=True, 
        readingOrder=1 if t['direction'] == 'rtl' else 2
    )
    
    # הוספת הסגנון לחוברת העבודה
    if 'header_style' not in wb.named_styles:
        wb.add_named_style(header_style)
    
    # סגנון תא רגיל
    cell_style = NamedStyle(name='cell_style')
    cell_style.font = Font(name='Arial', size=11)
    cell_style.border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    cell_style.alignment = Alignment(
        horizontal=t['alignment'],  # משתמש בערך המתאים לשפה - 'right' לעברית או 'left' לאנגלית
        vertical='center', 
        wrap_text=True, 
        readingOrder=1 if t['direction'] == 'rtl' else 2
    )
    
    # הוספת הסגנון לחוברת העבודה
    if 'cell_style' not in wb.named_styles:
        wb.add_named_style(cell_style)
    
    # סגנון לתא כותרת בדף הוראות
    instruction_header_style = NamedStyle(name='instruction_header_style')
    instruction_header_style.font = Font(name='Arial', size=14, bold=True)
    instruction_header_style.border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='medium', color='000000')
    )
    instruction_header_style.fill = PatternFill(fill_type='solid', start_color='4472C4', end_color='4472C4')
    instruction_header_style.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    instruction_header_style.alignment = Alignment(
        horizontal=t['alignment'],  # משתמש בערך המתאים לשפה - 'right' לעברית או 'left' לאנגלית
        vertical='center', 
        wrap_text=True, 
        readingOrder=1 if t['direction'] == 'rtl' else 2
    )
    
    # הוספת הסגנון לחוברת העבודה
    if 'instruction_header_style' not in wb.named_styles:
        wb.add_named_style(instruction_header_style)
    
    # סגנון לתא רגיל בדף הוראות
    instruction_cell_style = NamedStyle(name='instruction_cell_style')
    instruction_cell_style.font = Font(name='Arial', size=11)
    instruction_cell_style.border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    instruction_cell_style.alignment = Alignment(
        horizontal=t['alignment'],  # משתמש בערך המתאים לשפה - 'right' לעברית או 'left' לאנגלית
        vertical='center', 
        wrap_text=True, 
        readingOrder=1 if t['direction'] == 'rtl' else 2
    )
    
    # הוספת הסגנון לחוברת העבודה
    if 'instruction_cell_style' not in wb.named_styles:
        wb.add_named_style(instruction_cell_style)
        
    # סגנון כותרות משנה בדף הנחיות עם צבעים ספציפיים
    section_colors = {
        'main': '4472C4',  # צבע כחול לכותרת ראשית
        'settings': 'FF0000',  # אדום - תואם את גליון ההגדרות
        'main_categories': '70AD47',  # ירוק כהה - תואם את גליון קטגוריות ראשיות
        'sub_categories': '5B9BD5',  # כחול - תואם את גליון קטגוריות משנה
        'colors': 'C00000',  # אדום כהה - תואם את גליון צבעים
        'sizes': 'ED7D31',  # כתום - תואם את גליון מידות
        'materials': 'FFC000',  # צהוב-כתום - תואם את גליון חומרים
        'suppliers': '7030A0',  # סגול כהה - תואם את גליון ספקים
        'products': '00B050',  # ירוק בהיר - תואם את גליון מוצרים
        'notes': '808080',  # אפור - להערות והוראות כלליות
    }
    
    # יצירת סגנונות לכל סוג כותרת משנה בדף הנחיות
    for section, color in section_colors.items():
        style_name = f'section_{section}_style'
        section_style = NamedStyle(name=style_name)
        
        # שינוי צבע הטקסט בהתאם לבהירות הרקע
        text_color = 'FFFFFF'  # ברירת מחדל - לבן
        if section in ['materials', 'sizes']:  # צבעים בהירים
            text_color = '000000'  # שחור לרקעים בהירים
            
        section_style.font = Font(name='Arial', size=12, bold=True, color=text_color)
        section_style.fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
        section_style.border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )
        section_style.alignment = Alignment(
            horizontal=t['alignment'],  # משתמש בערך המתאים לשפה - 'right' לעברית או 'left' לאנגלית
            vertical='center', 
            wrap_text=True, 
            readingOrder=1 if t['direction'] == 'rtl' else 2
        )
        if style_name not in wb.named_styles:
            wb.add_named_style(section_style)
    
    # סגנון לכותרות בגליון הגדרות
    settings_header_style = NamedStyle(name='settings_header_style')
    settings_header_style.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    settings_header_style.fill = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
    settings_header_style.border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    settings_header_style.alignment = Alignment(
        horizontal=t['alignment'],  # משתמש בערך המתאים לשפה - 'right' לעברית או 'left' לאנגלית
        vertical='center', 
        wrap_text=True, 
        readingOrder=1 if t['direction'] == 'rtl' else 2
    )
    if 'settings_header_style' not in wb.named_styles:
        wb.add_named_style(settings_header_style)
        
    # סגנון לתאי נתונים בגליון הגדרות
    settings_value_style = NamedStyle(name='settings_value_style')
    settings_value_style.font = Font(name='Arial', size=11, bold=False)
    settings_value_style.border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    settings_value_style.alignment = Alignment(
        horizontal=t['alignment'],  # משתמש בערך המתאים לשפה - 'right' לעברית או 'left' לאנגלית
        vertical='center', 
        wrap_text=True, 
        readingOrder=1 if t['direction'] == 'rtl' else 2
    )
    if 'settings_value_style' not in wb.named_styles:
        wb.add_named_style(settings_value_style)

    # === גיליון הנחיות ===
    instructions_sheet = wb.active
    instructions_sheet.title = t['sheets']['instructions']
    instructions_sheet.sheet_view.rightToLeft = (t['direction'] == 'rtl')
    instructions_sheet.sheet_properties.tabColor = "0000FF"  # כחול
    
    # --------- שינויים בהטמעת הלוגו ----------
    # הוספת שורה ריקה בתחילת הגיליון עבור הלוגו
    instructions_sheet.insert_rows(1)
    instructions_sheet.row_dimensions[1].height = 120  # גובה שורה ללוגו - מוגדל ל-120 במקום 80
    
    # הגדרת רקע לבן לשורת הלוגו
    for col in range(1, 3):  # עבור עמודות A ו-B
        cell = instructions_sheet.cell(row=1, column=col)
        cell.fill = PatternFill(fill_type='solid', start_color='FFFFFF', end_color='FFFFFF')
        # נעילת התא כדי שלא יהיה ניתן לערוך אותו
        cell.protection = Protection(locked=True)
    
    # הוספת הלוגו
    logo_path = os.path.join('public', 'images', 'placeholder-logo.png')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 300  # רוחב הלוגו בפיקסלים - מוגדל ל-300 במקום 200
        img.height = 110  # גובה הלוגו בפיקסלים - מוגדל ל-110 במקום 80
        
        # מיקום מדויק של התמונה בתא B1
        # האנכור בתוך התא מבטיח שהתמונה תישאר בתא גם אם מזיזים שורות או עמודות
        
        # הגדרת מיקום מדויק בתוך התא (אפס פיקסלים מהקצה השמאלי והעליון של התא)
        marker = AnchorMarker(col=1, colOff=0, row=0, rowOff=0)  # תא B1
        
        # יצירת עיגון חדש מסוג OneCellAnchor (מעוגן לתא אחד)
        img.anchor = \ B1\
        
        # הגדרת התמונה כמעוגנת לתא וננעלת למניעת הזזה
        img.anchor.editAs = 'oneCell'  # מעוגן לתא
        
        # הוספת התמונה לגיליון
        instructions_sheet.add_image(img)
        
        # נעילת התמונה בתוספת הגדרות נוספות בהגנת הגיליון
        # הגנה זו תחול רק לאחר הפעלת הגנת הגיליון, שנעשית בתחתית הקוד
        instructions_sheet.protection.objects = True  # מגן על אובייקטים בגיליון
        instructions_sheet.protection.scenarios = True
        
        # הגדרת תא בעמודה A כתא ריק כדי לשמור על סדר טוב
        empty_cell = instructions_sheet.cell(row=1, column=1, value="")
    
    # הוספת שורה נוספת ריקה אחרי הלוגו לפני הכותרת
    instructions_sheet.insert_rows(2)
    instructions_sheet.row_dimensions[2].height = 10  # שורת רווח קטנה
    
    # עדכון רוחב עמודות - עמודה A לאינדקס/מספור, עמודה B לתוכן
    instructions_sheet.column_dimensions['A'].width = 6  # עמודה צרה יותר למספרים
    instructions_sheet.column_dimensions['B'].width = 94  # עמודה רחבה לתוכן

    # מיפוי בין שורות לסגנונות - מעודכן לפי החלוקה הנכונה
    section_mapping = [
        (3, 'instruction_header_style'),  # כותרת ראשית - עכשיו בשורה 3 (אחרי הלוגו והרווח)
        (5, 'section_main_style'),  # גרסה
        (6, 'section_main_style'),  # תאריך עדכון אחרון
        (7, 'section_main_style'),  # כל הזכויות שמורות
        (9, 'section_main_style'),  # מבנה הקובץ
        (10, 'section_main_style'),  # הקובץ מכיל את הגיליונות הבאים
        (11, 'section_main_style'),  # 1. הנחיות
        (12, 'section_settings_style'),  # 2. הגדרות
        (13, 'section_main_categories_style'),  # 3. קטגוריות ראשיות
        (14, 'section_sub_categories_style'),  # 4. קטגוריות משנה
        (15, 'section_colors_style'),  # 5. צבעים
        (16, 'section_sizes_style'),  # 6. מידות
        (17, 'section_materials_style'),  # 7. חומרים
        (18, 'section_suppliers_style'),  # 8. ספקים
        (19, 'section_products_style'),  # 9. מוצרים
        (21, 'section_notes_style'),  # הנחיות כלליות
        (22, 'section_notes_style'),  # 1. יש למלא את הנתונים
        (23, 'section_notes_style'),  # א. קטגוריות ראשיות
        (24, 'section_notes_style'),  # ב. קטגוריות משנה
        (25, 'section_notes_style'),  # ג. צבעים
        (26, 'section_notes_style'),  # ד. מידות
        (27, 'section_notes_style'),  # ה. חומרים
        (28, 'section_notes_style'),  # ו. ספקים
        (29, 'section_notes_style'),  # ז. מוצרים
        (31, 'section_notes_style'),  # 2. עמודות הקוד והמק"ט
        (32, 'section_notes_style'),  # 3. בשדות עם רשימה
        (33, 'section_notes_style'),  # 4. אין לשנות את מבנה הקובץ
        (34, 'section_notes_style'),  # 5. ניתן להוסיף שורות נתונים
        (36, 'section_notes_style'),  # פירוט השדות בכל גיליון
        (38, 'section_main_categories_style'),  # קטגוריות ראשיות
        (39, 'section_notes_style'),  # קוד: נוצר אוטומטית
        (40, 'section_notes_style'),  # שם: שם הקטגוריה
        (41, 'section_notes_style'),  # תיאור: תיאור מפורט
        (42, 'section_notes_style'),  # סטטוס: פעיל/לא פעיל
        (43, 'section_notes_style'),  # פעולה: עדכון/מחיקה/הוספה
        (45, 'section_sub_categories_style'),  # קטגוריות משנה
        (46, 'section_notes_style'),  # קוד: נוצר אוטומטית
        (47, 'section_notes_style'),  # שם: שם תת-הקטגוריה
        (48, 'section_notes_style'),  # תיאור: תיאור מפורט
        (49, 'section_notes_style'),  # קטגוריה ראשית: בחירה מרשימה
        (50, 'section_notes_style'),  # סטטוס: פעיל/לא פעיל
        (51, 'section_notes_style'),  # פעולה: עדכון/מחיקה/הוספה
        (53, 'section_colors_style'),  # צבעים
        (54, 'section_notes_style'),  # קוד: נוצר אוטומטית
        (55, 'section_notes_style'),  # שם: שם הצבע
        (56, 'section_notes_style'),  # תיאור: תיאור מפורט
        (57, 'section_notes_style'),  # סטטוס: פעיל/לא פעיל
        (58, 'section_notes_style'),  # פעולה: עדכון/מחיקה/הוספה
        (60, 'section_sizes_style'),  # מידות
        (61, 'section_notes_style'),  # קוד: נוצר אוטומטית
        (62, 'section_notes_style'),  # שם: המידה
        (63, 'section_notes_style'),  # תיאור: פירוט מדויק
        (64, 'section_notes_style'),  # סטטוס: פעיל/לא פעיל
        (65, 'section_notes_style'),  # פעולה: עדכון/מחיקה/הוספה
        (67, 'section_materials_style'),  # חומרים
        (68, 'section_notes_style'),  # קוד: נוצר אוטומטית
        (69, 'section_notes_style'),  # שם: שם החומר
        (70, 'section_notes_style'),  # תיאור: מפרט טכני
        (71, 'section_notes_style'),  # סטטוס: פעיל/לא פעיל
        (72, 'section_notes_style'),  # פעולה: עדכון/מחיקה/הוספה
        (74, 'section_suppliers_style'),  # ספקים
        (75, 'section_notes_style'),  # קוד: נוצר אוטומטית
        (76, 'section_notes_style'),  # שם: שם החברה
        (77, 'section_notes_style'),  # איש קשר: שם איש
        (78, 'section_notes_style'),  # אימייל: כתובת דוא"ל
        (79, 'section_notes_style'),  # טלפון: מספר טלפון
        (80, 'section_notes_style'),  # כתובת: כתובת מלאה
        (81, 'section_notes_style'),  # סטטוס: פעיל/לא פעיל
        (82, 'section_notes_style'),  # פעולה: עדכון/מחיקה/הוספה
        (84, 'section_products_style'),  # מוצרים
        (85, 'section_notes_style'),  # מק"ט: נוצר אוטומטית
        (86, 'section_notes_style'),  # שם: שם המוצר
        (87, 'section_notes_style'),  # תיאור: תיאור מפורט
        (88, 'section_notes_style'),  # קטגוריה ראשית: בחירה מרשימה
        (89, 'section_notes_style'),  # קטגוריית משנה: בחירה מרשימה
        (90, 'section_notes_style'),  # ספק: בחירה מרשימה
        (91, 'section_notes_style'),  # צבעים: בחירה מרשימה
        (92, 'section_notes_style'),  # מידות: בחירה מרשימה
        (93, 'section_notes_style'),  # חומרים: בחירה מרשימה
        (94, 'section_notes_style'),  # כמות באריזה: מספר יחידות
        (95, 'section_notes_style'),  # הוראות אריזה: הנחיות מיוחדות
        (96, 'section_notes_style'),  # כמות בקרטון: מספר יחידות
        (97, 'section_notes_style'),  # מחיר ליחידה: מחיר ב-USD
        (98, 'section_notes_style'),  # סטטוס: פעיל/לא פעיל
        (99, 'section_notes_style'),  # פעולה: עדכון/מחיקה/הוספה
        (101, 'section_notes_style'),  # הערות חשובות
        (102, 'section_notes_style'),  # 1. יש לוודא שכל השדות
        (103, 'section_notes_style'),  # 2. בעת הוספת מוצר חדש
        (104, 'section_notes_style'),  # 3. מחירים יש להזין
        (105, 'section_notes_style'),  # 4. בכל שינוי יש לציין
        (106, 'section_notes_style'),  # 5. אין למחוק שורות
        (108, 'section_notes_style'),  # תמיכה
        (109, 'section_notes_style'),  # במקרה של תקלה
        (110, 'section_notes_style'),  # support@piordersystem.com
    ]
    
    # פונקציה משופרת לפיצול תוכן ההנחיות למספרים ותוכן
    def split_instruction_content(row_idx, text):
        """
        מפצל את הטקסט למספור ותוכן
        row_idx: אינדקס השורה
        text: הטקסט המלא
        מחזיר: (number, content) - המספר והתוכן המפוצלים
        """
        import re
        
        # פיצול לטקסט הכולל מספרים/אותיות ותוכן
        if t['direction'] == 'rtl':  # עברית
            # בדיקה אם יש מספור בסוף שורה (למשל "הנחיות - גיליון נוכחי .1")
            rtl_number_pattern = re.compile(r'(.*?)\s+\.(\d+)$')
            rtl_letter_pattern = re.compile(r'(.*?)\s+\.([א-ת])$')
            rtl_dash_pattern = re.compile(r'(.*?)\s+-$')
            
            # חיפוש מספר בסוף
            match = rtl_number_pattern.match(text)
            if match:
                return '.' + match.group(2), match.group(1)
            
            # חיפוש אות בסוף
            match = rtl_letter_pattern.match(text)
            if match:
                return '.' + match.group(2), match.group(1)
            
            # חיפוש מקף בסוף
            match = rtl_dash_pattern.match(text)
            if match:
                return '-', match.group(1)
            
            # מקרים ספציפיים לדפוסי הטקסט שבטבלה
            if text.endswith(' -'):
                parts = text.split(' -', 1)
                return '-', parts[0]
            
            # חיפוש פסיק עם מספר (במקום נקודה) כמו "תיאור: תיאור מפורט של הקטגוריה ,"
            rtl_comma_number_pattern = re.compile(r'(.*?)\s+,(\d+)$')
            match = rtl_comma_number_pattern.match(text)
            if match:
                return '.' + match.group(2), match.group(1)
                
            return '', text
        else:  # אנגלית
            # בדיקת דפוסים שונים באנגלית
            ltr_number_pattern = re.compile(r'^(\d+)[\.\)] (.*)')
            ltr_letter_pattern = re.compile(r'^([a-g])[\.\)] (.*)')
            ltr_dash_pattern = re.compile(r'^- (.*)')
            
            # חיפוש מספר בתחילה
            match = ltr_number_pattern.match(text)
            if match:
                return match.group(1) + '.', match.group(2)
            
            # חיפוש אות בתחילה
            match = ltr_letter_pattern.match(text)
            if match:
                return match.group(1) + '.', match.group(2)
            
            # חיפוש מקף בתחילה
            match = ltr_dash_pattern.match(text)
            if match:
                return '-', match.group(1)
                
            return '', text
    
    # לפני המילוי של הטבלה, נעבור על כל השורות ונפצל את הטקסט לשתי עמודות
    preprocessed_instructions = []
    for row_idx, instruction in enumerate(t['instructions_text'], 1):
        text = instruction[0]
        number, content = split_instruction_content(row_idx, text)
        
        # מקרים מיוחדים שלא נתפסים בפונקציה הכללית
        # כאשר יש שני נתונים בשורה אחת כמו "קוד: נוצר אוטומטית (CAT001, CAT002, ...) -"
        if ":" in content and not number and t['direction'] == 'rtl':
            # אם יש נקודתיים וזו שפה עברית, וזה לא נקלט כבר
            number = '-'
            
        preprocessed_instructions.append((number, content))
    
    # מילוי הטבלה עם הנתונים המעובדים
    for row_idx, (number, content) in enumerate(preprocessed_instructions, 1):
        # הגדרת סגנון לשורה מסוימת
        matching_section = next((style for r, style in section_mapping if r == row_idx), None)
        style_to_apply = 'instruction_header_style' if row_idx == 1 else (matching_section or 'instruction_cell_style')
        
        # הוספת הערכים לתאים עם הסגנון המתאים
        num_cell = instructions_sheet.cell(row=row_idx, column=1, value=number)
        content_cell = instructions_sheet.cell(row=row_idx, column=2, value=content)
        
        num_cell.style = style_to_apply
        content_cell.style = style_to_apply
    
    # הגדרת הגנה על הגיליון
    instructions_sheet.protection.sheet = True
    instructions_sheet.protection.password = SHEET_PASSWORD
    instructions_sheet.protection.enable()

    # === גיליון הגדרות ===
    settings_sheet = wb.create_sheet(t['sheets']['settings'])
    settings_sheet.sheet_view.rightToLeft = (t['direction'] == 'rtl')
    settings_sheet.sheet_properties.tabColor = "FF0000"  # אדום
    
    # הגדרות מק"ט בעברית ובאנגלית
    sku_settings = {
        'he': [
            ['הגדרות מק"ט', ''],
            ['קידומת', 'HY'],
            ['כמות תווים', '6'],
            ['דוגמה', '=B2&REPT("0",B3-LEN(B2)-1)&"1"']  # נוסחה נכונה שמוסיפה מספר נכון של אפסים
        ],
        'en': [
            ['SKU Settings', ''],
            ['Prefix', 'HY'],
            ['Number of Characters', '6'],
            ['Example', '=B2&REPT("0",B3-LEN(B2)-1)&"1"']  # נוסחה נכונה שמוסיפה מספר נכון של אפסים
        ]
    }[lang]
    
    for row_idx, (setting, value) in enumerate(sku_settings, 1):
        settings_sheet.cell(row=row_idx, column=1, value=setting)
        settings_sheet.cell(row=row_idx, column=2, value=value)
        
        # יישום סגנונות
        if row_idx == 1:
            settings_sheet.cell(row=row_idx, column=1).style = 'settings_header_style'
            settings_sheet.cell(row=row_idx, column=2).style = 'settings_header_style'
        else:
            settings_sheet.cell(row=row_idx, column=1).style = 'settings_value_style'
            settings_sheet.cell(row=row_idx, column=2).style = 'settings_value_style'
    
    # הוספת הערה לתאים הפתוחים לעריכה
    prefix_comment = Comment('ניתן לשנות את קידומת המק"ט לפי הצורך', 'System')
    settings_sheet.cell(row=2, column=2).comment = prefix_comment
    
    digits_comment = Comment('ניתן לשנות את מספר התווים הכולל של המק"ט', 'System')
    settings_sheet.cell(row=3, column=2).comment = digits_comment
    
    # עדכון הגנות הגליון עם סיסמה
    for sheet in wb.worksheets:
        if sheet.title != t['sheets']['instructions']:
            sheet.protection.sheet = True
            sheet.protection.password = SHEET_PASSWORD
            sheet.protection.enable()
            
            # הגדרת הרשאות ספציפיות
            sheet.protection.formatCells = False
            sheet.protection.formatColumns = False
            sheet.protection.formatRows = False
            sheet.protection.insertColumns = False
            sheet.protection.insertRows = True
            sheet.protection.deleteColumns = False
            sheet.protection.deleteRows = True
            sheet.protection.sort = True
            sheet.protection.autoFilter = True
            
            # נעילת כותרות בלבד (לא עמודת קוד/מק"ט)
            for row in range(1, sheet.max_row + 1):
                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row, column=col)
                    if row == 1:  # רק כותרות נעולות
                        cell.protection = Protection(locked=True)
                    else:
                        cell.protection = Protection(locked=False)

    # יישום הגנה מחודשת לגיליון הגדרות לאחר הלולאה הכללית
    settings_sheet = wb[t['sheets']['settings']]
    # נעילת כל התאים בגיליון הגדרות מלבד התאים של קידומת המק"ט וכמות התווים
    for row in range(1, 5):
        for col in range(1, 3):
            cell = settings_sheet.cell(row=row, column=col)
            # נעילת כל התאים מלבד B2 ו-B3 (קידומת וכמות תווים)
            if row == 2 and col == 2:  # תא B2 - קידומת
                cell.protection = Protection(locked=False)
            elif row == 3 and col == 2:  # תא B3 - כמות תווים
                cell.protection = Protection(locked=False)
            else:
                cell.protection = Protection(locked=True)
    
    # הפעלת הגנה על הגיליון
    settings_sheet.protection.sheet = True
    settings_sheet.protection.password = SHEET_PASSWORD
    settings_sheet.protection.enable()
    
    # הגדרת הרשאות ספציפיות
    settings_sheet.protection.formatCells = False
    settings_sheet.protection.formatColumns = False
    settings_sheet.protection.formatRows = False
    settings_sheet.protection.insertColumns = False
    settings_sheet.protection.insertRows = False
    settings_sheet.protection.deleteColumns = False
    settings_sheet.protection.deleteRows = False
    settings_sheet.protection.sort = False
    settings_sheet.protection.autoFilter = False

    settings_sheet.column_dimensions['A'].width = 20
    settings_sheet.column_dimensions['B'].width = 20

    # עדכון הטקסט בהנחיות - תיקון המספור
    t['instructions_text'] = [
        [text[0].replace(' 1.', '1.').replace(' 2.', '2.').replace(' 3.', '3.').replace(' 4.', '4.').replace(' 5.', '5.')] 
        if any(num in text[0] for num in ['1.', '2.', '3.', '4.', '5.'])
        else [text[0]]
        for text in t['instructions_text']
    ]

    # === יצירת כל הגיליונות ===
    sheets_config = {
        'main_categories': {
            'headers': [t['headers']['code'], t['headers']['name'], t['headers']['description'], t['headers']['status'], t['headers']['action']],
            'widths': [15, 30, 40, 15, 15],
            'color': '70AD47'  # ירוק כהה - תואם לגיליון ההנחיות
        },
        'sub_categories': {
            'headers': [t['headers']['code'], t['headers']['name'], t['headers']['description'], t['headers']['main_category'], t['headers']['status'], t['headers']['action']],
            'widths': [15, 30, 40, 30, 15, 15],
            'color': '5B9BD5'  # כחול - תואם לגיליון ההנחיות
        },
        'colors': {
            'headers': [t['headers']['code'], t['headers']['name'], t['headers']['description'], t['headers']['status'], t['headers']['action']],
            'widths': [15, 30, 40, 15, 15],
            'color': 'C00000'  # אדום כהה - תואם לגיליון ההנחיות
        },
        'sizes': {
            'headers': [t['headers']['code'], t['headers']['name'], t['headers']['description'], t['headers']['status'], t['headers']['action']],
            'widths': [15, 30, 40, 15, 15],
            'color': 'ED7D31'  # כתום - תואם לגיליון ההנחיות
        },
        'materials': {
            'headers': [t['headers']['code'], t['headers']['name'], t['headers']['description'], t['headers']['status'], t['headers']['action']],
            'widths': [15, 30, 40, 15, 15],
            'color': 'FFC000'  # צהוב-כתום - תואם לגיליון ההנחיות
        },
        'suppliers': {
            'headers': [t['headers']['code'], t['headers']['name'], t['headers']['contact'], t['headers']['email'], t['headers']['phone'], t['headers']['address'], t['headers']['status'], t['headers']['action']],
            'widths': [15, 30, 30, 40, 20, 50, 15, 15],
            'color': '7030A0'  # סגול כהה - תואם לגיליון ההנחיות
        },
        'products': {
            'headers': [
                t['headers']['sku'], t['headers']['name'], t['headers']['description'],
                t['headers']['main_category'], t['headers']['sub_category'], t['headers']['supplier'],
                t['headers']['colors'], t['headers']['sizes'], t['headers']['materials'],
                t['headers']['units_per_pack'], t['headers']['packing_instructions'],
                t['headers']['units_per_carton'], t['headers']['price_per_unit'],
                t['headers']['status'], t['headers']['action']
            ],
            'widths': [15, 30, 40, 20, 20, 20, 30, 30, 20, 15, 40, 15, 15, 15, 15],
            'color': '00B050'  # ירוק בהיר - תואם לגיליון ההנחיות
        }
    }

    for sheet_name, config in sheets_config.items():
        sheet = wb.create_sheet(t['sheets'][sheet_name])
        sheet.sheet_view.rightToLeft = (t['direction'] == 'rtl')
        sheet.sheet_properties.tabColor = config['color']
        
        # הגדרת גובה שורת כותרת
        sheet.row_dimensions[1].height = 30  # גובה שורת כותרת
        
        # הגדרת כותרות ורוחב עמודות
        for col, (header, width) in enumerate(zip(config['headers'], config['widths']), 1):
            # הגדרת ערך התא
            cell = sheet.cell(row=1, column=col, value=header)
            # יישום סגנון כותרת
            cell.style = 'header_style'
            # הגדרת רוחב עמודה
            sheet.column_dimensions[get_column_letter(col)].width = width
        
        # הוספת סגנון לכל התאים
        for row in range(2, 1001):
            for col in range(1, len(config['headers']) + 1):
                cell = sheet.cell(row=row, column=col)
                cell.style = 'cell_style'

        # הוספת נוסחאות לקודים אוטומטיים
        prefix = {
            'main_categories': 'CAT',
            'sub_categories': None,  # מטופל בנפרד
            'colors': 'COL',
            'sizes': 'SIZ',
            'materials': 'MAT',
            'suppliers': 'SUP',
            'products': None  # מטופל בנפרד - נלקח מגיליון הגדרות
        }[sheet_name]
        
        if prefix:
            for row in range(2, 1001):
                cell = sheet.cell(row=row, column=1)
                cell.protection = Protection(locked=True)
                cell.value = generate_code_formula(prefix, row)
        elif sheet_name == 'sub_categories':
            for row in range(2, 1001):
                cell = sheet.cell(row=row, column=1)
                cell.protection = Protection(locked=True)
                cell.value = f'=IF(AND(B{row}<>"",D{row}<>""),INDEX(\'{t["sheets"]["main_categories"]}\'!$A$2:$A$1000,MATCH(D{row},\'{t["sheets"]["main_categories"]}\'!$B$2:$B$1000,0))&"-"&TEXT(COUNTIF(D$2:D{row},D{row}),"000"),"")'
        elif sheet_name == 'products':
            # נוסחה למק"ט בעמודה הראשונה
            for row in range(2, 1001):
                cell = sheet.cell(row=row, column=1)
                cell.protection = Protection(locked=False)  # התא לא נעול - מאפשר הזנה ידנית
                
                # שימוש באותה נוסחה בדיוק כמו בדוגמה בגיליון הגדרות
                cell.value = f'=IF(B{row}<>"",\'{t["sheets"]["settings"]}\'!B2&REPT("0",\'{t["sheets"]["settings"]}\'!B3-LEN(\'{t["sheets"]["settings"]}\'!B2)-1)&(ROW()-1),"")'

            # הוספת הערה לעמודת המק"ט
            comment = Comment('ניתן להזין מק"ט ידנית או להשאיר את הנוסחה האוטומטית.\nהזנת ערך ידנית תדרוס את הנוסחה האוטומטית.', 'System')
            sheet['A1'].comment = comment

    # עדכון ולידציות
    for sheet_name in sheets_config:
        sheet = wb[t['sheets'][sheet_name]]
        status_col = len(sheet[1]) - 1  # עמודה לפני אחרונה
        action_col = len(sheet[1])  # עמודה אחרונה
        
        # ולידציה לסטטוס
        status_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(t["validations"]["status"])}"',
            allow_blank=False,
            showErrorMessage=True,
            errorTitle=t['validations']['error_title'],
            error=t['validations']['error_message']
        )
        sheet.add_data_validation(status_validation)
        status_validation.add(f'{get_column_letter(status_col)}2:{get_column_letter(status_col)}1000')

        # ולידציה לפעולה
        action_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(t["validations"]["action"])}"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle=t['validations']['error_title'],
            error=t['validations']['error_message']
        )
        sheet.add_data_validation(action_validation)
        action_validation.add(f'{get_column_letter(action_col)}2:{get_column_letter(action_col)}1000')

    # עדכון ולידציות נוספות בגיליון מוצרים
    products_sheet = wb[t['sheets']['products']]
    products_validations = {
        'main_category': {'col': 4, 'source': t['sheets']['main_categories']},
        'sub_category': {'col': 5, 'source': t['sheets']['sub_categories']},
        'supplier': {'col': 6, 'source': t['sheets']['suppliers']},
        'colors': {'col': 7, 'source': t['sheets']['colors']},
        'sizes': {'col': 8, 'source': t['sheets']['sizes']},
        'materials': {'col': 9, 'source': t['sheets']['materials']}
    }

    for field, config in products_validations.items():
        validation = DataValidation(
            type="list",
            formula1=f'=OFFSET(\'{config["source"]}\'!$B$2,0,0,COUNTA(\'{config["source"]}\'!B:B)-1)',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle=t['validations']['error_title'],
            error=t['validations']['error_message']
        )
        products_sheet.add_data_validation(validation)
        validation.add(f'{get_column_letter(config["col"])}2:{get_column_letter(config["col"])}1000')

    # הוספת ולידציה לקטגוריה ראשית בגיליון קטגוריות משנה
    sub_categories_sheet = wb[t['sheets']['sub_categories']]
    main_category_validation = DataValidation(
        type="list",
        formula1=f'=OFFSET(\'{t["sheets"]["main_categories"]}\'!$B$2,0,0,COUNTA(\'{t["sheets"]["main_categories"]}\'!B:B)-1)',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle=t['validations']['error_title'],
        error=t['validations']['error_message']
    )
    sub_categories_sheet.add_data_validation(main_category_validation)
    main_category_validation.add('D2:D1000')

    # שמירת הקובץ
    wb.save(output_file)
    print(f"Template created successfully at {output_file}!")

    # לנוחות, נשמור גם גרסה רגילה ללא VBA
    wb.save(output_file)
    print(f"Standard template also saved at {output_file}!")

if __name__ == "__main__":
    # יצירת גרסה בעברית
    create_master_template(lang='he')
    
    # יצירת גרסה באנגלית
    create_master_template(lang='en') 




