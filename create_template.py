import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection, NamedStyle
from openpyxl.styles.colors import Color
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
import os
from openpyxl.formatting.rule import FormulaRule
import re

def generate_code_formula(prefix, start_row):
    """יצירת נוסחה לקוד אוטומטי שמתעלמת משורות ריקות"""
    return f'=IF(B{start_row}<>"","{prefix}"&TEXT(COUNTIF(B$2:B{start_row},"<>"),"000"),"")'

def generate_subcat_code_formula(start_row, sheet_name):
    """יצירת נוסחה לקוד קטגוריית משנה שמתעלמת משורות ריקות"""
    return f'=IF(AND(B{start_row}<>"",D{start_row}<>""),INDEX(\'{sheet_name}\'!$A$2:$A$1000,MATCH(D{start_row},\'{sheet_name}\'!$B$2:$B$1000,0))&"-"&TEXT(COUNTIF(D$2:D{start_row},"<>"),"000"),"")'

def create_master_template(lang='he'):
    """
    יצירת תבנית אקסל
    lang: 'he' לעברית, 'en' לאנגלית
    """
    VERSION = "1.0.0"  # גרסת התבנית
    CURRENT_DATE = "30/03/2025"  # תאריך עדכון אחרון
    SHEET_PASSWORD = "FlexiPI@2025#System"  # סיסמה להגנת הגליון
    SUPPORT_EMAIL = "support@flexipi.com"  # כתובת אימייל לתמיכה
    
    # מילון תרגומים
    translations = {
        'he': {
            'direction': 'rtl',
            'alignment': 'right',
            'sheets': {
                'main_categories': 'קטגוריות ראשיות',
                'sub_categories': 'קטגוריות משנה',
                'colors': 'צבעים',
                'sizes': 'מידות',
                'materials': 'חומרים',
                'suppliers': 'ספקים',
                'products': 'מוצרים',
                'instructions': 'הנחיות',
                'settings': 'הגדרות'
            },
            'headers': {
                'code': 'קוד',
                'name': 'שם',
                'description': 'תיאור',
                'status': 'סטטוס',
                'action': 'פעולה',
                'main_category': 'קטגוריה ראשית',
                'sub_category': 'קטגוריית משנה',
                'contact': 'איש קשר',
                'email': 'אימייל',
                'phone': 'טלפון',
                'address': 'כתובת',
                'sku': 'מק"ט',
                'supplier': 'ספק',
                'colors': 'צבעים',
                'sizes': 'מידות',
                'materials': 'חומרים',
                'units_per_pack': 'כמות באריזה',
                'packing_instructions': 'הוראות אריזה',
                'units_per_carton': 'כמות בקרטון',
                'price_per_unit': 'מחיר ליחידה',
                'start_code': 'קוד התחלתי'
            },
            'validations': {
                'status': ['פעיל', 'לא פעיל'],
                'action': ['עדכון', 'מחיקה', 'הוספה'],
                'error_title': 'שגיאה',
                'error_message': 'יש לבחור ערך מהרשימה בלבד'
            },
            'instructions_text': [
                [''],  # Row for logo
                [''],  # Spacing row
                [''],  # Row for header
                ['הנחיות שימוש בקובץ טמפלייט לניהול הזמנות FlexiPI'],
                [''],
                [f'גרסה: {VERSION}'],
                [f'תאריך עדכון אחרון: {CURRENT_DATE}'],
                ['כל הזכויות שמורות © FlexiPI'],
                [''],
                ['מבנה הקובץ:'],
                ['הקובץ מכיל את הגיליונות הבאים:'],
                ['הנחיות - הגיליון הנוכחי .1'],
                ['הגדרות - הגדרות מערכת כגון מבנה מק"ט .2'],
                ['קטגוריות ראשיות - ניהול קטגוריות ראשיות של מוצרים .3'],
                ['קטגוריות משנה - ניהול תתי-קטגוריות .4'],
                ['צבעים - ניהול צבעים זמינים .5'],
                ['מידות - ניהול מידות זמינות .6'],
                ['חומרים - ניהול חומרים זמינים .7'],
                ['ספקים - ניהול פרטי ספקים .8'],
                ['מוצרים - ניהול פריטי המלאי .9'],
                [''],
                ['הנחיות כלליות:'],
                ['יש למלא את הנתונים בסדר הבא: .1'],
                ['קטגוריות ראשיות .א'],
                ['קטגוריות משנה .ב'],
                ['צבעים .ג'],
                ['מידות .ד'],
                ['חומרים .ה'],
                ['ספקים .ו'],
                ['מוצרים .ז'],
                [''],
                ['עמודות הקוד והמק"ט נעולות ומתמלאות אוטומטית (מק"ט בלבד ניתן להזין גם ידנית) .2'],
                ['בשדות עם רשימה נפתחת יש לבחור מהרשימה בלבד .3'],
                ['אין לשנות את מבנה הקובץ או להוסיף/למחוק עמודות .4'],
                ['ניתן להוסיף שורות נתונים ככל שנדרש .5'],
                [''],
                ['פירוט השדות בכל גיליון:'],
                [''],
                ['קטגוריות ראשיות:'],
                ['קוד: נוצר אוטומטית (CAT001, CAT002, ...) -'],
                ['שם: שם הקטגוריה (לדוגמה: "כלי מטבח", "ריהוט") -'],
                ['תיאור: תיאור מפורט של הקטגוריה -'],
                ['סטטוס: פעיל/לא פעיל -'],
                ['פעולה: עדכון/מחיקה/הוספה -'],
                [''],
                ['קטגוריות משנה:'],
                ['קוד: נוצר אוטומטית (CAT001-001, CAT001-002, ...) -'],
                ['שם: שם תת-הקטגוריה (לדוגמה: "סכו"ם", "כיסאות") -'],
                ['תיאור: תיאור מפורט של תת-הקטגוריה -'],
                ['קטגוריה ראשית: בחירה מרשימה נפתחת -'],
                ['סטטוס: פעיל/לא פעיל -'],
                ['פעולה: עדכון/מחיקה/הוספה -'],
                [''],
                ['צבעים:'],
                ['קוד: נוצר אוטומטית (COL001, COL002, ...) -'],
                ['שם: שם הצבע (לדוגמה: "אדום", "כחול") -'],
                ['תיאור: תיאור מפורט או קוד צבע -'],
                ['סטטוס: פעיל/לא פעיל -'],
                ['פעולה: עדכון/מחיקה/הוספה -'],
                [''],
                ['מידות:'],
                ['קוד: נוצר אוטומטית (SIZ001, SIZ002, ...) -'],
                ['שם: המידה (לדוגמה: "S", "M", "L", "XL") -'],
                ['תיאור: פירוט מדויק של המידה -'],
                ['סטטוס: פעיל/לא פעיל -'],
                ['פעולה: עדכון/מחיקה/הוספה -'],
                [''],
                ['חומרים:'],
                ['קוד: נוצר אוטומטית (MAT001, MAT002, ...) -'],
                ['שם: שם החומר (לדוגמה: "עץ אלון", "פלסטיק") -'],
                ['תיאור: מפרט טכני של החומר -'],
                ['סטטוס: פעיל/לא פעיל -'],
                ['פעולה: עדכון/מחיקה/הוספה -'],
                [''],
                ['ספקים:'],
                ['קוד: נוצר אוטומטית (SUP001, SUP002, ...) -'],
                ['שם: שם החברה -'],
                ['איש קשר: שם איש הקשר -'],
                ['אימייל: כתובת דוא"ל -'],
                ['טלפון: מספר טלפון -'],
                ['כתובת: כתובת מלאה -'],
                ['קוד התחלתי: מספר התחלתי לקידוד מוצרי הספק (אופציונלי) -'],
                ['סטטוס: פעיל/לא פעיל -'],
                ['פעולה: עדכון/מחיקה/הוספה -'],
                [''],
                ['מוצרים:'],
                ['מק"ט: נוצר אוטומטית או ידנית -'],
                ['שם: שם המוצר -'],
                ['תיאור: תיאור מפורט של המוצר -'],
                ['קטגוריה ראשית: בחירה מרשימה -'],
                ['קטגוריית משנה: בחירה מרשימה -'],
                ['ספק: בחירה מרשימה -'],
                ['צבעים: בחירה מרשימה -'],
                ['מידות: בחירה מרשימה -'],
                ['חומרים: בחירה מרשימה -'],
                ['כמות באריזה: מספר יחידות -'],
                ['הוראות אריזה: הנחיות מיוחדות -'],
                ['כמות בקרטון: מספר יחידות -'],
                ['מחיר ליחידה: מחיר ב-USD -'],
                ['סטטוס: פעיל/לא פעיל -'],
                ['פעולה: עדכון/מחיקה/הוספה -'],
                [''],
                ['הערות חשובות:'],
                ['יש לוודא שכל השדות החובה מלאים .1'],
                ['בעת הוספת מוצר חדש, יש לוודא שכל הערכים הנבחרים (קטגוריות, צבעים וכו\') קיימים ופעילים .2'],
                ['מחירים יש להזין במספרים בלבד, ללא סימנים מיוחדים .3'],
                ['בכל שינוי יש לציין את סוג הפעולה בעמודת "פעולה" .4'],
                ['אין למחוק שורות קיימות, אלא רק לסמן אותן כ"לא פעיל" .5'],
                ['קוד התחלתי בספקים: מאפשר להתחיל את מספור המוצרים ממספר מסוים עבור כל ספק .6'],
                ['חשוב: הזנת מק"ט ידני תשבור את רצף המספור האוטומטי עבור אותו ספק .7'],
                [''],
                ['תמיכה:'],
                ['במקרה של תקלה או שאלה, ניתן לפנות לתמיכה הטכנית:'],
                [SUPPORT_EMAIL]
            ]
        },
        'en': {
            'direction': 'ltr',
            'alignment': 'left',
            'sheets': {
                'main_categories': 'Main Categories',
                'sub_categories': 'Sub Categories',
                'colors': 'Colors',
                'sizes': 'Sizes',
                'materials': 'Materials',
                'suppliers': 'Suppliers',
                'products': 'Products',
                'instructions': 'Instructions',
                'settings': 'Settings'
            },
            'headers': {
                'code': 'Code',
                'name': 'Name',
                'description': 'Description',
                'status': 'Status',
                'action': 'Action',
                'main_category': 'Main Category',
                'sub_category': 'Sub Category',
                'contact': 'Contact',
                'email': 'Email',
                'phone': 'Phone',
                'address': 'Address',
                'sku': 'SKU',
                'supplier': 'Supplier',
                'colors': 'Colors',
                'sizes': 'Sizes',
                'materials': 'Materials',
                'units_per_pack': 'Units Per Pack',
                'packing_instructions': 'Packing Instructions',
                'units_per_carton': 'Units Per Carton',
                'price_per_unit': 'Price Per Unit',
                'start_code': 'Start Code'
            },
            'validations': {
                'status': ['Active', 'Inactive'],
                'action': ['Update', 'Delete', 'Add'],
                'error_title': 'Error',
                'error_message': 'Please select a value from the list'
            },
            'instructions_text': [
                [''],  # Row for logo
                [''],  # Spacing row
                [''],  # Row for header
                ['FlexiPI Template Usage Instructions'],
                [''],
                [f'Version: {VERSION}'],
                [f'Last Updated: {CURRENT_DATE}'],
                ['All Rights Reserved © FlexiPI'],
                [''],
                ['File Structure:'],
                ['The file contains the following sheets:'],
                ['1. Instructions - Current sheet'],
                ['2. Settings - System settings like SKU structure'],
                ['3. Main Categories - Manage main product categories'],
                ['4. Sub Categories - Manage sub-categories'],
                ['5. Colors - Manage available colors'],
                ['6. Sizes - Manage available sizes'],
                ['7. Materials - Manage available materials'],
                ['8. Suppliers - Manage supplier details'],
                ['9. Products - Manage inventory items'],
                [''],
                ['General Instructions:'],
                ['1. Fill in the data in the following order:'],
                ['   a. Main Categories'],
                ['   b. Sub Categories'],
                ['   c. Colors'],
                ['   d. Sizes'],
                ['   e. Materials'],
                ['   f. Suppliers'],
                ['   g. Products'],
                [''],
                ['Code and SKU columns are locked and auto-filled (SKU can also be entered manually) .2'],
                ['In fields with dropdown lists, select only from the list .3'],
                ['Do not modify the file structure or add/delete columns .4'],
                ['You can add data rows as needed .5'],
                [''],
                ['Field Details for Each Sheet:'],
                [''],
                ['Main Categories:'],
                ['- Code: Auto-generated (CAT001, CAT002, ...)'],
                ['- Name: Category name (e.g., "Kitchen Utensils", "Furniture")'],
                ['- Description: Detailed category description'],
                ['- Status: Active/Inactive'],
                ['- Action: Update/Delete/Add'],
                [''],
                ['Sub Categories:'],
                ['- Code: Auto-generated (CAT001-001, CAT001-002, ...)'],
                ['- Name: Sub-category name (e.g., "Cutlery", "Chairs")'],
                ['- Description: Detailed sub-category description'],
                ['- Main Category: Select from dropdown'],
                ['- Status: Active/Inactive'],
                ['- Action: Update/Delete/Add'],
                [''],
                ['Colors:'],
                ['- Code: Auto-generated (COL001, COL002, ...)'],
                ['- Name: Color name (e.g., "Red", "Blue")'],
                ['- Description: Detailed description or color code'],
                ['- Status: Active/Inactive'],
                ['- Action: Update/Delete/Add'],
                [''],
                ['Sizes:'],
                ['- Code: Auto-generated (SIZ001, SIZ002, ...)'],
                ['- Name: Size (e.g., "S", "M", "L", "XL")'],
                ['- Description: Exact size specifications'],
                ['- Status: Active/Inactive'],
                ['- Action: Update/Delete/Add'],
                [''],
                ['Materials:'],
                ['- Code: Auto-generated (MAT001, MAT002, ...)'],
                ['- Name: Material name (e.g., "Oak Wood", "Plastic")'],
                ['- Description: Technical specifications'],
                ['- Status: Active/Inactive'],
                ['- Action: Update/Delete/Add'],
                [''],
                ['Suppliers:'],
                ['- Code: Auto-generated (SUP001, SUP002, ...)'],
                ['- Name: Company name'],
                ['- Contact: Contact person name'],
                ['- Email: Email address'],
                ['- Phone: Phone number'],
                ['- Address: Full address'],
                ['- Status: Active/Inactive'],
                ['- Action: Update/Delete/Add'],
                [''],
                ['Products:'],
                ['- SKU: Editable or auto-generated based on settings'],
                ['- Name: Product name'],
                ['- Description: Detailed product description'],
                ['- Main Category: Select from list'],
                ['- Sub Category: Select from list'],
                ['- Supplier: Select from list'],
                ['- Colors: Select from list'],
                ['- Sizes: Select from list'],
                ['- Materials: Select from list'],
                ['- Units Per Pack: Number of units'],
                ['- Packing Instructions: Special instructions'],
                ['- Units Per Carton: Number of units'],
                ['- Price Per Unit: Price in USD'],
                ['- Status: Active/Inactive'],
                ['- Action: Update/Delete/Add'],
                [''],
                ['Important Notes:'],
                ['1. Ensure all required fields are filled'],
                ['2. When adding a new product, verify all selected values (categories, colors, etc.) exist and are active'],
                ['3. Enter prices as numbers only, without special characters'],
                ['4. Specify the action type in the "Action" column for any changes'],
                ['5. Do not delete existing rows, mark them as "Inactive" instead'],
                ['6. Start Code in Suppliers: Allows starting product numbering from a specific number for each supplier'],
                ['7. Important: Entering a manual SKU will break the automatic numbering sequence for that supplier'],
                [''],
                ['Support:'],
                ['For technical support or questions, contact:'],
                [SUPPORT_EMAIL]
            ]
        }
    }

    t = translations[lang]
    
    # יצירת תיקיית היעד
    output_dir = os.path.join('public', 'templates')
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f'FlexiPI_Template_v{VERSION}_{CURRENT_DATE.replace("/", "_")}_{lang}.xlsx')

    # יצירת חוברת עבודה
    wb = openpyxl.Workbook()
    
    # סגנונות כלליים
    header_style = NamedStyle(name='header_style')
    header_style.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_style.fill = PatternFill(fill_type='solid', start_color='2E7D32', end_color='2E7D32')
    header_style.border = Border(
        left=Side(border_style='medium', color='000000'),
        right=Side(border_style='medium', color='000000'),
        top=Side(border_style='medium', color='000000'),
        bottom=Side(border_style='medium', color='000000')
    )
    header_style.alignment = Alignment(
        horizontal='center', 
        vertical='center', 
        wrap_text=True, 
        readingOrder=1 if t['direction'] == 'rtl' else 2
    )
    if 'header_style' not in wb.named_styles:
        wb.add_named_style(header_style)
    
    cell_style = NamedStyle(name='cell_style')
    cell_style.font = Font(name='Arial', size=11)
    cell_style.border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    cell_style.alignment = Alignment(
        horizontal=t['alignment'],
        vertical='center', 
        wrap_text=True, 
        readingOrder=1 if t['direction'] == 'rtl' else 2
    )
    if 'cell_style' not in wb.named_styles:
        wb.add_named_style(cell_style)
    
    instruction_header_style = NamedStyle(name='instruction_header_style')
    instruction_header_style.font = Font(name='Arial', size=14, bold=True, color='FFFFFF')
    instruction_header_style.fill = PatternFill(fill_type='solid', start_color='4472C4', end_color='4472C4')
    instruction_header_style.border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='medium', color='000000')
    )
    instruction_header_style.alignment = Alignment(
        horizontal=t['alignment'],
        vertical='center', 
        wrap_text=True, 
        readingOrder=1 if t['direction'] == 'rtl' else 2
    )
    if 'instruction_header_style' not in wb.named_styles:
        wb.add_named_style(instruction_header_style)
    
    instruction_cell_style = NamedStyle(name='instruction_cell_style')
    instruction_cell_style.font = Font(name='Arial', size=11)
    instruction_cell_style.border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    instruction_cell_style.alignment = Alignment(
        horizontal=t['alignment'],
        vertical='center', 
        wrap_text=True, 
        readingOrder=1 if t['direction'] == 'rtl' else 2
    )
    if 'instruction_cell_style' not in wb.named_styles:
        wb.add_named_style(instruction_cell_style)
        
    section_colors = {
        'main': '4472C4',
        'settings': 'FF0000',
        'main_categories': '70AD47',
        'sub_categories': '5B9BD5',
        'colors': 'C00000',
        'sizes': 'ED7D31',
        'materials': 'FFC000',
        'suppliers': '7030A0',
        'products': '00B050',
        'notes': '808080'
    }
    
    for section, color in section_colors.items():
        style_name = f'section_{section}_style'
        section_style = NamedStyle(name=style_name)
        text_color = 'FFFFFF' if section not in ['materials', 'sizes'] else '000000'
        section_style.font = Font(name='Arial', size=12, bold=True, color=text_color)
        section_style.fill = PatternFill(fill_type='solid', start_color=color, end_color=color)
        section_style.border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )
        section_style.alignment = Alignment(
            horizontal=t['alignment'],
            vertical='center', 
            wrap_text=True, 
            readingOrder=1 if t['direction'] == 'rtl' else 2
        )
        if style_name not in wb.named_styles:
            wb.add_named_style(section_style)
    
    settings_header_style = NamedStyle(name='settings_header_style')
    settings_header_style.font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
    settings_header_style.fill = PatternFill(fill_type='solid', start_color='FF0000', end_color='FF0000')
    settings_header_style.border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    settings_header_style.alignment = Alignment(
        horizontal=t['alignment'],
        vertical='center', 
        wrap_text=True, 
        readingOrder=1 if t['direction'] == 'rtl' else 2
    )
    if 'settings_header_style' not in wb.named_styles:
        wb.add_named_style(settings_header_style)
        
    settings_value_style = NamedStyle(name='settings_value_style')
    settings_value_style.font = Font(name='Arial', size=11, bold=False)
    settings_value_style.border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )
    settings_value_style.alignment = Alignment(
        horizontal=t['alignment'],
        vertical='center', 
        wrap_text=True, 
        readingOrder=1 if t['direction'] == 'rtl' else 2
    )
    if 'settings_value_style' not in wb.named_styles:
        wb.add_named_style(settings_value_style)

    # גיליון הנחיות
    instructions_sheet = wb.active
    instructions_sheet.title = t['sheets']['instructions']
    instructions_sheet.sheet_view.rightToLeft = (t['direction'] == 'rtl')
    instructions_sheet.sheet_properties.tabColor = "0000FF"
    
    instructions_sheet.insert_rows(1, 3)
    instructions_sheet.row_dimensions[1].height = 100
    instructions_sheet.row_dimensions[2].height = 20
    instructions_sheet.row_dimensions[3].height = 30
    
    logo_path = os.path.join('public', 'images', 'placeholder-logo.jpg')
    if os.path.exists(logo_path):
        img = Image(logo_path)
        img.width = 799
        img.height = 134
        instructions_sheet.merge_cells('A1:B1')
        merged_cell = instructions_sheet['A1']
        merged_cell.alignment = Alignment(horizontal='center', vertical='center')
        merged_cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
        merged_cell.border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )
        merged_cell.font = Font(name='Arial', size=11)
        img.anchor = 'A1'
        instructions_sheet.add_image(img)
        merged_cell.protection = Protection(locked=True)

    instructions_sheet.column_dimensions['A'].width = 6
    instructions_sheet.column_dimensions['B'].width = 94
    
    for row_idx in [1, 2]:
        for cell in instructions_sheet[row_idx]:
            cell.fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            cell.border = Border(
                left=Side(border_style='thin', color='000000'),
                right=Side(border_style='thin', color='000000'),
                top=Side(border_style='thin', color='000000'),
                bottom=Side(border_style='thin', color='000000')
            )
    
    section_mapping = [
        (4, 'instruction_header_style'),  # כותרת ראשית
        (6, 'section_main_style'),        # גרסה
        (7, 'section_main_style'),        # תאריך עדכון
        (8, 'section_main_style'),        # זכויות יוצרים
        (10, 'section_main_style'),       # מבנה הקובץ
        (11, 'section_main_style'),       # הקובץ מכיל
        (12, 'section_main_style'),       # הנחיות
        (13, 'section_settings_style'),   # הגדרות
        (14, 'section_main_categories_style'),  # קטגוריות ראשיות
        (15, 'section_sub_categories_style'),   # קטגוריות משנה
        (16, 'section_colors_style'),     # צבעים
        (17, 'section_sizes_style'),      # מידות
        (18, 'section_materials_style'),  # חומרים
        (19, 'section_suppliers_style'),  # ספקים
        (20, 'section_products_style'),   # מוצרים
        (22, 'section_notes_style'),      # הנחיות כלליות
        (23, 'section_notes_style'),      # יש למלא את הנתונים
        (24, 'section_notes_style'),      # קטגוריות ראשיות
        (25, 'section_notes_style'),      # קטגוריות משנה
        (26, 'section_notes_style'),      # צבעים
        (27, 'section_notes_style'),      # מידות
        (28, 'section_notes_style'),      # חומרים
        (29, 'section_notes_style'),      # ספקים
        (30, 'section_notes_style'),      # מוצרים
        (32, 'section_notes_style'),      # עמודות הקוד והמק"ט
        (33, 'section_notes_style'),      # בשדות עם רשימה
        (34, 'section_notes_style'),      # אין לשנות את מבנה הקובץ
        (35, 'section_notes_style'),      # ניתן להוסיף שורות
        (37, 'section_notes_style'),      # פירוט השדות
        (39, 'section_main_categories_style'),  # קטגוריות ראשיות
        (40, 'section_notes_style'),      # קוד
        (41, 'section_notes_style'),      # שם
        (42, 'section_notes_style'),      # תיאור
        (43, 'section_notes_style'),      # סטטוס
        (44, 'section_notes_style'),      # פעולה
        (46, 'section_sub_categories_style'),   # קטגוריות משנה
        (47, 'section_notes_style'),      # קוד
        (48, 'section_notes_style'),      # שם
        (49, 'section_notes_style'),      # תיאור
        (50, 'section_notes_style'),      # קטגוריה ראשית
        (51, 'section_notes_style'),      # סטטוס
        (52, 'section_notes_style'),      # פעולה
        (54, 'section_colors_style'),     # צבעים
        (55, 'section_notes_style'),      # קוד
        (56, 'section_notes_style'),      # שם
        (57, 'section_notes_style'),      # תיאור
        (58, 'section_notes_style'),      # סטטוס
        (59, 'section_notes_style'),      # פעולה
        (61, 'section_sizes_style'),      # מידות
        (62, 'section_notes_style'),      # קוד
        (63, 'section_notes_style'),      # שם
        (64, 'section_notes_style'),      # תיאור
        (65, 'section_notes_style'),      # סטטוס
        (66, 'section_notes_style'),      # פעולה
        (68, 'section_materials_style'),  # חומרים
        (69, 'section_notes_style'),      # קוד
        (70, 'section_notes_style'),      # שם
        (71, 'section_notes_style'),      # תיאור
        (72, 'section_notes_style'),      # סטטוס
        (73, 'section_notes_style'),      # פעולה
        (75, 'section_suppliers_style'),  # ספקים
        (76, 'section_notes_style'),      # קוד
        (77, 'section_notes_style'),      # שם
        (78, 'section_notes_style'),      # איש קשר
        (79, 'section_notes_style'),      # אימייל
        (80, 'section_notes_style'),      # טלפון
        (81, 'section_notes_style'),      # כתובת
        (82, 'section_notes_style'),      # קוד התחלתי
        (83, 'section_notes_style'),      # סטטוס
        (84, 'section_notes_style'),      # פעולה
        (86, 'section_products_style'),   # מוצרים
        (87, 'section_notes_style'),      # מק"ט
        (88, 'section_notes_style'),      # שם
        (89, 'section_notes_style'),      # תיאור
        (90, 'section_notes_style'),      # קטגוריה ראשית
        (91, 'section_notes_style'),      # קטגוריית משנה
        (92, 'section_notes_style'),      # ספק
        (93, 'section_notes_style'),      # צבעים
        (94, 'section_notes_style'),      # מידות
        (95, 'section_notes_style'),      # חומרים
        (96, 'section_notes_style'),      # כמות באריזה
        (97, 'section_notes_style'),      # הוראות אריזה
        (98, 'section_notes_style'),      # כמות בקרטון
        (99, 'section_notes_style'),      # מחיר ליחידה
        (100, 'section_notes_style'),     # סטטוס
        (101, 'section_notes_style'),     # פעולה
        (103, 'section_notes_style'),     # הערות חשובות
        (104, 'section_notes_style'),     # יש לוודא שכל השדות
        (105, 'section_notes_style'),     # בעת הוספת מוצר חדש
        (106, 'section_notes_style'),     # מחירים יש להזין
        (107, 'section_notes_style'),     # בכל שינוי יש לציין
        (108, 'section_notes_style'),     # אין למחוק שורות
        (109, 'section_notes_style'),     # קוד התחלתי בספקים
        (110, 'section_notes_style'),     # חשוב: הזנת מק"ט ידני
        (112, 'section_notes_style'),     # תמיכה
        (113, 'section_notes_style'),     # במקרה של תקלה
        (114, 'section_notes_style')      # אימייל תמיכה
    ]
    
    def split_instruction_content(row_idx, text):
        if t['direction'] == 'rtl':
            rtl_number_pattern = re.compile(r'(.*?)\s+\.(\d+)$')
            rtl_letter_pattern = re.compile(r'(.*?)\s+\.([א-ת])$')
            rtl_dash_pattern = re.compile(r'(.*?)\s+-$')
            match = rtl_number_pattern.match(text)
            if match:
                return '.' + match.group(2), match.group(1)
            match = rtl_letter_pattern.match(text)
            if match:
                return '.' + match.group(2), match.group(1)
            match = rtl_dash_pattern.match(text)
            if match:
                return '-', match.group(1)
            if text.endswith(' -'):
                parts = text.split(' -', 1)
                return '-', parts[0]
            return '', text
        else:
            ltr_number_pattern = re.compile(r'^(\d+)[\.\)] (.*)')
            ltr_letter_pattern = re.compile(r'^([a-g])[\.\)] (.*)')
            ltr_dash_pattern = re.compile(r'^- (.*)')
            match = ltr_number_pattern.match(text)
            if match:
                return match.group(1) + '.', match.group(2)
            match = ltr_letter_pattern.match(text)
            if match:
                return match.group(1) + '.', match.group(2)
            match = ltr_dash_pattern.match(text)
            if match:
                return '-', match.group(1)
            return '', text
    
    preprocessed_instructions = []
    for row_idx, instruction in enumerate(t['instructions_text'], 1):
        text = instruction[0]
        number, content = split_instruction_content(row_idx, text)
        if ":" in content and not number and t['direction'] == 'rtl':
            number = '-'
        preprocessed_instructions.append((number, content))
    
    for row_idx, (number, content) in enumerate(preprocessed_instructions, 1):
        if row_idx == 1:
            continue
        num_cell = instructions_sheet.cell(row=row_idx, column=1, value=number)
        content_cell = instructions_sheet.cell(row=row_idx, column=2, value=content)
        style_to_apply = next((style for r, style in section_mapping if r == row_idx), 'instruction_cell_style')
        num_cell.style = style_to_apply
        content_cell.style = style_to_apply
    
    for row in range(1, instructions_sheet.max_row + 1):
        for col in range(1, instructions_sheet.max_column + 1):
            instructions_sheet.cell(row=row, column=col).protection = Protection(locked=True)
    
    instructions_sheet.protection.sheet = True
    instructions_sheet.protection.enable()

    # עדכון הגנות הגליון עם סיסמה
    for sheet in wb.worksheets:
        if sheet.title != t['sheets']['instructions']:
            sheet.protection.sheet = True
            sheet.protection.password = SHEET_PASSWORD
            sheet.protection.enable()
            
            # הגדרת הרשאות ספציפיות
            sheet.protection.formatCells = False
            sheet.protection.formatColumns = False
            sheet.protection.formatRows = False
            sheet.protection.insertColumns = False
            sheet.protection.insertRows = True
            sheet.protection.deleteColumns = False
            sheet.protection.deleteRows = True
            sheet.protection.sort = True
            sheet.protection.autoFilter = True
            sheet.protection.scenarios = True  # מאפשר תרחישים
            sheet.protection.objects = True    # מגן על אובייקטים
            sheet.protection.pivotTables = True  # מגן על טבלאות ציר
            sheet.protection.selectLockedCells = True  # מאפשר בחירת תאים נעולים
            sheet.protection.selectUnlockedCells = True  # מאפשר בחירת תאים לא נעולים

    # גיליון הגדרות
    settings_sheet = wb.create_sheet(t['sheets']['settings'])
    settings_sheet.sheet_view.rightToLeft = (t['direction'] == 'rtl')
    settings_sheet.sheet_properties.tabColor = "FF0000"
    
    sku_settings = {
        'he': [
            ['הגדרות מק"ט', ''],
            ['קידומת', 'HY'],
            ['כמות תווים', '6'],
            ['דוגמה', '=B2&REPT("0",B3-LEN(B2)-1)&"1"']
        ],
        'en': [
            ['SKU Settings', ''],
            ['Prefix', 'HY'],
            ['Number of Characters', '6'],
            ['Example', '=B2&REPT("0",B3-LEN(B2)-1)&"1"']
        ]
    }[lang]
    
    for row_idx, (setting, value) in enumerate(sku_settings, 1):
        settings_sheet.cell(row=row_idx, column=1, value=setting).style = 'settings_header_style' if row_idx == 1 else 'settings_value_style'
        settings_sheet.cell(row=row_idx, column=2, value=value).style = 'settings_header_style' if row_idx == 1 else 'settings_value_style'
    
    settings_sheet.cell(row=2, column=2).comment = Comment('ניתן לשנות את קידומת המק"ט לפי הצורך', 'System')
    settings_sheet.cell(row=3, column=2).comment = Comment('ניתן לשנות את מספר התווים הכולל של המק"ט', 'System')
    
    for row in range(1, 5):
        for col in range(1, 3):
            cell = settings_sheet.cell(row=row, column=col)
            cell.protection = Protection(locked=False if (row == 2 or row == 3) and col == 2 else True)
    
    settings_sheet.protection.sheet = True
    settings_sheet.protection.enable()
    
    settings_sheet.column_dimensions['A'].width = 20
    settings_sheet.column_dimensions['B'].width = 20

    # יצירת כל הגיליונות
    sheets_config = {
        'main_categories': {
            'headers': [t['headers']['code'], t['headers']['name'], t['headers']['description'], t['headers']['status'], t['headers']['action']],
            'widths': [15, 30, 40, 15, 15],
            'color': '70AD47',
            'lock_code': True  # נעילת עמודת הקוד
        },
        'sub_categories': {
            'headers': [t['headers']['code'], t['headers']['name'], t['headers']['description'], t['headers']['main_category'], t['headers']['status'], t['headers']['action']],
            'widths': [15, 30, 40, 30, 15, 15],
            'color': '5B9BD5',
            'lock_code': True  # נעילת עמודת הקוד
        },
        'colors': {
            'headers': [t['headers']['code'], t['headers']['name'], t['headers']['description'], t['headers']['status'], t['headers']['action']],
            'widths': [15, 30, 40, 15, 15],
            'color': 'C00000',
            'lock_code': True  # נעילת עמודת הקוד
        },
        'sizes': {
            'headers': [t['headers']['code'], t['headers']['name'], t['headers']['description'], t['headers']['status'], t['headers']['action']],
            'widths': [15, 30, 40, 15, 15],
            'color': 'ED7D31',
            'lock_code': True  # נעילת עמודת הקוד
        },
        'materials': {
            'headers': [t['headers']['code'], t['headers']['name'], t['headers']['description'], t['headers']['status'], t['headers']['action']],
            'widths': [15, 30, 40, 15, 15],
            'color': 'FFC000',
            'lock_code': True  # נעילת עמודת הקוד
        },
        'suppliers': {
            'headers': [t['headers']['code'], t['headers']['name'], t['headers']['contact'], t['headers']['email'], t['headers']['phone'], t['headers']['address'], t['headers']['start_code'], t['headers']['status'], t['headers']['action']],
            'widths': [15, 30, 30, 40, 20, 50, 15, 15, 15],
            'color': '7030A0',
            'lock_code': True  # נעילת עמודת הקוד
        },
        'products': {
            'headers': [
                t['headers']['sku'], t['headers']['name'], t['headers']['description'],
                t['headers']['main_category'], t['headers']['sub_category'], t['headers']['supplier'],
                t['headers']['colors'], t['headers']['sizes'], t['headers']['materials'],
                t['headers']['units_per_pack'], t['headers']['packing_instructions'],
                t['headers']['units_per_carton'], t['headers']['price_per_unit'],
                t['headers']['status'], t['headers']['action']
            ],
            'widths': [15, 30, 40, 20, 20, 20, 30, 30, 20, 15, 40, 15, 15, 15, 15],
            'color': '00B050',
            'lock_code': False  # עמודת המק"ט פתוחה לעריכה
        }
    }

    for sheet_name, config in sheets_config.items():
        sheet = wb.create_sheet(t['sheets'][sheet_name])
        sheet.sheet_view.rightToLeft = (t['direction'] == 'rtl')
        sheet.sheet_properties.tabColor = config['color']
        sheet.row_dimensions[1].height = 30
        
        # הגדרת כותרות
        for col, (header, width) in enumerate(zip(config['headers'], config['widths']), 1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.style = 'header_style'
            sheet.column_dimensions[get_column_letter(col)].width = width
            cell.protection = Protection(locked=True)
        
        # הגדרת תאים ונוסחאות
        for row in range(2, 1001):
            for col in range(1, len(config['headers']) + 1):
                cell = sheet.cell(row=row, column=col)
                cell.style = 'cell_style'
                # נעילה לפי lock_code: עמודה A נעולה אם lock_code=True, פתוחה אם False
                if col == 1:
                    cell.protection = Protection(locked=config['lock_code'])
                else:
                    cell.protection = Protection(locked=False)

    # הוספת נוסחאות לקודים אוטומטיים
        prefix = {
            'main_categories': 'CAT',
            'sub_categories': None,
            'colors': 'COL',
            'sizes': 'SIZ',
            'materials': 'MAT',
            'suppliers': 'SUP',
            'products': None
        }[sheet_name]
        
        if prefix:
            for row in range(2, 1001):
                sheet.cell(row=row, column=1).value = generate_code_formula(prefix, row)
        elif sheet_name == 'sub_categories':
            for row in range(2, 1001):
                sheet.cell(row=row, column=1).value = generate_subcat_code_formula(row, t['sheets']['main_categories'])
        elif sheet_name == 'products':
            for row in range(2, 1001):
                sheet.cell(row=row, column=1).value = (
                    f'=IF(AND(B{row}<>"",F{row}<>""),'  # בדיקה שיש שם מוצר וספק
                    f'  \'{t["sheets"]["settings"]}\'!B2&'  # קידומת מההגדרות
                    f'  TEXT('
                    f'    IF(VLOOKUP(F{row},\'{t["sheets"]["suppliers"]}\'!B:G,6,FALSE)="",'  # אם אין קוד התחלתי לספק
                    f'      COUNTIFS(F$2:F{row},F{row}),'  # ספירת מוצרים של אותו ספק
                    f'      VALUE(VLOOKUP(F{row},\'{t["sheets"]["suppliers"]}\'!B:G,6,FALSE)) + COUNTIFS(F$2:F{row},F{row}) - 1'  # קוד התחלתי + ספירה
                    f'    )'
                    f'  ,"0000")'
                    f',"")'
                )
        
        # הגדרת הגנות הגליון
        sheet.protection.sheet = True
        sheet.protection.enable()

    # עדכון ולידציות
    for sheet_name in sheets_config:
        sheet = wb[t['sheets'][sheet_name]]
        status_col = len(sheet[1]) - 1
        action_col = len(sheet[1])
        
        status_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(t["validations"]["status"])}"',
            allow_blank=False,
            showErrorMessage=True,
            errorTitle=t['validations']['error_title'],
            error=t['validations']['error_message']
        )
        sheet.add_data_validation(status_validation)
        status_validation.add(f'{get_column_letter(status_col)}2:{get_column_letter(status_col)}1000')

        action_validation = DataValidation(
            type="list",
            formula1=f'"{",".join(t["validations"]["action"])}"',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle=t['validations']['error_title'],
            error=t['validations']['error_message']
        )
        sheet.add_data_validation(action_validation)
        action_validation.add(f'{get_column_letter(action_col)}2:{get_column_letter(action_col)}1000')

    products_sheet = wb[t['sheets']['products']]
    products_validations = {
        'main_category': {'col': 4, 'source': t['sheets']['main_categories']},
        'sub_category': {'col': 5, 'source': t['sheets']['sub_categories']},
        'supplier': {'col': 6, 'source': t['sheets']['suppliers']},
        'colors': {'col': 7, 'source': t['sheets']['colors']},
        'sizes': {'col': 8, 'source': t['sheets']['sizes']},
        'materials': {'col': 9, 'source': t['sheets']['materials']}
    }

    for field, config in products_validations.items():
        validation = DataValidation(
            type="list",
            formula1=f'=OFFSET(\'{config["source"]}\'!$B$2,0,0,COUNTA(\'{config["source"]}\'!B:B)-1)',
            allow_blank=True,
            showErrorMessage=True,
            errorTitle=t['validations']['error_title'],
            error=t['validations']['error_message']
        )
        products_sheet.add_data_validation(validation)
        validation.add(f'{get_column_letter(config["col"])}2:{get_column_letter(config["col"])}1000')

    sub_categories_sheet = wb[t['sheets']['sub_categories']]
    main_category_validation = DataValidation(
        type="list",
        formula1=f'=OFFSET(\'{t["sheets"]["main_categories"]}\'!$B$2,0,0,COUNTA(\'{t["sheets"]["main_categories"]}\'!B:B)-1)',
        allow_blank=True,
        showErrorMessage=True,
        errorTitle=t['validations']['error_title'],
        error=t['validations']['error_message']
    )
    sub_categories_sheet.add_data_validation(main_category_validation)
    main_category_validation.add('D2:D1000')

    # הגדרות עדכון אוטומטי של קישורים
    wb.keep_links = False  # לא לשמור קישורים
    wb.keep_vba = True    # שומר על מאקרו VBA אם קיים
    
    # שמירת הקובץ עם הגדרות אבטחה זהות
    wb_protection = openpyxl.workbook.protection.WorkbookProtection(
        workbookPassword=SHEET_PASSWORD,
        lockStructure=True,
        lockWindows=True
    )
    wb_protection.updateLinks = False  # מניעת עדכון קישורים אוטומטי
    wb.security = wb_protection
    
    # הגדרת תכונות נוספות לקובץ
    for sheet in wb.worksheets:
        sheet.sheet_state = 'visible'  # וידוא שכל הגליונות גלויים
        if hasattr(sheet, 'updateLinks'):
            sheet.updateLinks = False  # מניעת עדכון קישורים ברמת הגליון
    
    # הגדרות נוספות למניעת עדכון קישורים
    wb.properties.allowRefreshQuery = False  # מניעת רענון שאילתות
    if hasattr(wb, 'disable_links'):
        wb.disable_links = True  # מניעת קישורים חיצוניים
    if hasattr(wb, 'update_links'):
        wb.update_links = False  # מניעת עדכון קישורים
    
    # שמירת הקובץ עם הגדרות מיוחדות
    wb.save(output_file)
    
    print(f"Template created successfully at {output_file}!")

if __name__ == "__main__":
    create_master_template(lang='he')
    create_master_template(lang='en') 